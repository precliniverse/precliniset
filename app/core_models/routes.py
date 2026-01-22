# app/core_models/routes.py
import json
import os
from datetime import datetime

import pandas as pd
from flask import (abort, current_app, flash, redirect, render_template,
                   request, send_file, send_from_directory, session, url_for)
from flask_babel import lazy_gettext as _l
from flask_login import current_user, login_required
from flask_wtf.csrf import validate_csrf
from sqlalchemy import func
from werkzeug.utils import secure_filename
from sqlalchemy.orm.attributes import flag_modified

from app.utils.files import (dataframe_to_excel_bytes, read_excel_to_list,
                             validate_file_type)

from ..extensions import db
from ..forms import (AnimalModelBulkUploadForm, CoreModelForm,
                     ProtocolBulkUploadForm, ProtocolModelForm)
from ..helpers import generate_unique_name, generate_xlsx_template
from ..models import (Analyte, AnalyteDataType, AnimalModel,
                      AnimalModelAnalyteAssociation, ControlledMolecule, DataTable,
                      ExperimentalGroup, ImportPipeline, ProtocolAnalyteAssociation,
                      ProtocolAttachment, ProtocolModel, ProtocolMoleculeAssociation,
                      Severity)
from ..permissions import check_group_permission, user_has_permission
from ..services.tm_connector import TrainingManagerConnector
from . import core_models_bp


@core_models_bp.route('/bulk_upload_animal_model', methods=['GET', 'POST'])
@login_required
def bulk_upload_animal_model():
    if not user_has_permission(current_user, 'CoreModel', 'create', allow_any_team=True):
        flash(_l("You do not have permission to perform this action."), "danger")
        return redirect(url_for('core_models.manage_models'))
    
    form = AnimalModelBulkUploadForm()
    if form.validate_on_submit():
        file = form.file.data
        try:
            # Use utility to read Excel
            data_list, columns = read_excel_to_list(file, sheet_name="Animal Model Bulk Upload")
            df = pd.DataFrame(data_list) # Convert back to DF for grouping logic below

            required_cols = ["Animal Model Name", "Analyte Name"]
            if not all(col in columns for col in required_cols):
                flash(_l("The uploaded file is missing required columns. Please use the template."), "danger")
                return redirect(url_for('core_models.bulk_upload_animal_model'))

            models_to_create = {}
            analytes_to_create = {}
            conflicts = []
            models_to_skip = []

            for model_name, group in df.groupby('Animal Model Name'):
                existing_model = AnimalModel.query.filter(func.lower(AnimalModel.name) == func.lower(model_name)).first()
                if existing_model:
                    models_to_skip.append(model_name)
                    continue

                model_info = {"name": model_name, "analytes": []}
                
                for index, row in group.iterrows():
                    analyte_name = row['Analyte Name']
                    if not analyte_name:
                        continue

                    analyte_link_info = {
                        "name": analyte_name,
                        "default_value": row.get('Default Value'),
                        "is_metadata": str(row.get('Is Metadata (TRUE/FALSE)')).upper() == 'TRUE'
                    }

                    existing_analyte = Analyte.query.filter(func.lower(Analyte.name) == func.lower(analyte_name)).first()
                    if existing_analyte:
                        if pd.notna(row.get('New Analyte Unit')) and row.get('New Analyte Unit') != existing_analyte.unit:
                            conflicts.append(f"Analyte '{analyte_name}': Unit mismatch. File has '{row.get('New Analyte Unit')}', DB has '{existing_analyte.unit}'.")
                        if pd.notna(row.get('New Analyte Data Type')) and row.get('New Analyte Data Type') != existing_analyte.data_type.value:
                            conflicts.append(f"Analyte '{analyte_name}': Data type mismatch. File has '{row.get('New Analyte Data Type')}', DB has '{existing_analyte.data_type.value}'.")
                    elif analyte_name not in analytes_to_create:
                        new_analyte_info = {
                            "name": analyte_name,
                            "description": row.get('New Analyte Description'),
                            "unit": row.get('New Analyte Unit'),
                            "data_type": row.get('New Analyte Data Type'),
                            "allowed_values": row.get('New Analyte Allowed Values')
                        }
                        if not new_analyte_info['data_type'] or new_analyte_info['data_type'] not in [dt.value for dt in AnalyteDataType]:
                            conflicts.append(f"New Analyte '{analyte_name}': Invalid data type '{new_analyte_info['data_type']}'.")
                        if new_analyte_info['data_type'] == 'category' and not new_analyte_info['allowed_values']:
                            conflicts.append(f"New Analyte '{analyte_name}': 'Allowed Values' is required for category data type.")
                        analytes_to_create[analyte_name] = new_analyte_info

                    model_info["analytes"].append(analyte_link_info)
                models_to_create[model_name] = model_info

            session['bulk_upload_animal_model_data'] = {
                "models": list(models_to_create.values()),
                "new_analytes": list(analytes_to_create.values()),
                "conflicts": conflicts,
                "skipped_models": models_to_skip
            }
            
            return redirect(url_for('core_models.bulk_upload_animal_model_review'))

        except Exception as e:
            flash(_l("An error occurred while processing the file: %(error)s", error=str(e)), "danger")
            return redirect(url_for('core_models.bulk_upload_animal_model'))

    return render_template('core_models/bulk_upload_animal_model.html', form=form)

@core_models_bp.route('/bulk_upload_animal_model_review', methods=['GET'])
@login_required
def bulk_upload_animal_model_review():
    if not user_has_permission(current_user, 'CoreModel', 'create', allow_any_team=True):
        abort(403)
    
    upload_data = session.get('bulk_upload_animal_model_data', {})
    if not upload_data:
        flash(_l("No upload data found in session. Please start again."), "warning")
        return redirect(url_for('core_models.bulk_upload_animal_model'))

    return render_template('core_models/bulk_upload_animal_model_review.html',
                           models=upload_data.get('models', []),
                           new_analytes=upload_data.get('new_analytes', []),
                           conflicts=upload_data.get('conflicts', []))

@core_models_bp.route('/bulk_upload_animal_model_confirm', methods=['POST'])
@login_required
def bulk_upload_animal_model_confirm():
    if not user_has_permission(current_user, 'CoreModel', 'create', allow_any_team=True):
        abort(403)

    upload_data = session.get('bulk_upload_animal_model_data', {})
    if not upload_data or upload_data.get('conflicts'):
        flash(_l("Cannot confirm import due to conflicts or missing data."), "danger")
        return redirect(url_for('core_models.bulk_upload_animal_model_review'))

    try:
        new_analyte_objects = {}
        for analyte_data in upload_data.get('new_analytes', []):
            new_analyte = Analyte(
                name=analyte_data['name'],
                description=analyte_data['description'],
                unit=analyte_data['unit'],
                data_type=[dt for dt in AnalyteDataType if dt.value == analyte_data['data_type']],
                allowed_values=analyte_data['allowed_values'],
                creator_id=current_user.id
            )
            db.session.add(new_analyte)
            new_analyte_objects[analyte_data['name']] = new_analyte
        db.session.flush()

        for model_data in upload_data.get('models', []):
            new_model = AnimalModel(name=model_data['name'])
            db.session.add(new_model)
            db.session.flush()

            linked_analyte_ids = set()
            for analyte_link in model_data['analytes']:
                analyte_obj = new_analyte_objects.get(analyte_link['name']) or Analyte.query.filter(func.lower(Analyte.name) == func.lower(analyte_link['name'])).first()
                if analyte_obj:
                    linked_analyte_ids.add(analyte_obj.id)
            
            mandatory_analytes = Analyte.query.filter(Analyte.name.in_(['ID', 'Date of Birth'])).all()
            for man_analyte in mandatory_analytes:
                linked_analyte_ids.add(man_analyte.id)

            new_model.analytes = Analyte.query.filter(Analyte.id.in_(linked_analyte_ids)).all()

        db.session.commit()
        session.pop('bulk_upload_animal_model_data', None)
        flash(_l("Bulk import successful!"), "success")
        return redirect(url_for('core_models.manage_models'))

    except Exception as e:
        db.session.rollback()
        flash(_l("An error occurred during the final import: %(error)s", error=str(e)), "danger")
        return redirect(url_for('core_models.bulk_upload_animal_model_review'))

def generate_animal_model_export(models):
    all_model_data = []
    for model in models:
        associations = AnimalModelAnalyteAssociation.query.filter_by(animal_model_id=model.id).order_by(AnimalModelAnalyteAssociation.order).all()
        
        if not associations:
            all_model_data.append({
                "Animal Model Name": model.name, "Analyte Name": "",
                "Is Metadata (TRUE/FALSE)": "", "Default Value": "",
                "New Analyte Description": "", "New Analyte Unit": "",
                "New Analyte Data Type": "", "New Analyte Allowed Values": ""
            })
        else:
            for assoc in associations:
                analyte = assoc.analyte
                model_data = {
                    "Animal Model Name": model.name, "Analyte Name": analyte.name,
                    "Is Metadata (TRUE/FALSE)": str(analyte.is_metadata).upper(),
                    "Default Value": analyte.default_value,
                    "New Analyte Description": analyte.description, "New Analyte Unit": analyte.unit,
                    "New Analyte Data Type": analyte.data_type.value, "New Analyte Allowed Values": analyte.allowed_values
                }
                all_model_data.append(model_data)
    
    df = pd.DataFrame(all_model_data)
    return dataframe_to_excel_bytes(df, sheet_name="Animal Model Bulk Upload")

@core_models_bp.route('/download_selected_animal_models', methods=['POST'])
@login_required
def download_selected_animal_models():
    selected_ids = request.form.getlist('selected_models')
    if not selected_ids:
        flash(_l("No models selected for download."), "warning")
        return redirect(url_for('core_models.manage_models'))

    models = AnimalModel.query.filter(AnimalModel.id.in_(selected_ids)).all()
    if not models:
        flash(_l("Selected models not found."), "danger")
        return redirect(url_for('core_models.manage_models'))

    excel_file = generate_animal_model_export(models)
    filename = f"selected_animal_models_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return send_file(excel_file, mimetype=mimetype, as_attachment=True, download_name=filename)

@core_models_bp.route('/download_animal_model_template')
@login_required
def download_animal_model_template():
    # This uses openpyxl directly for formatting, so we keep it here or move to a template service
    try:
        from io import BytesIO

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Animal Model Bulk Upload"

        headers = [
            "Animal Model Name", "Analyte Name", "Is Metadata (TRUE/FALSE)", "Default Value",
            "New Analyte Description", "New Analyte Unit", "New Analyte Data Type", "New Analyte Allowed Values"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        model_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
        analyte_fill = PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid")
        new_analyte_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")

        for i, cell in enumerate(ws[1]):
            cell.font = header_font
            if i < 1:
                cell.fill = model_fill
            elif i < 4:
                cell.fill = analyte_fill
            else:
                cell.fill = new_analyte_fill

        instructions = [
            "", "Instructions:",
            "1. Each row represents one analyte linked to an animal model.",
            "2. To create one model with multiple analytes, repeat the 'Animal Model Name' in each row.",
            "3. 'Is Metadata' should be TRUE or FALSE.",
            "4. The 'New Analyte...' columns are only needed if the 'Analyte Name' does not already exist.",
            "5. 'New Analyte Data Type' must be one of: float, int, text, category, date.",
            "6. For 'category' data types, 'New Analyte Allowed Values' must be a semicolon-separated list (e.g., Value1;Value2;Value3)."
        ]
        for instruction in instructions:
            ws.append([instruction])

        example_data = [
            ("C57BL/6 Mouse", "Body Weight", "FALSE", "20.5", "Initial body weight", "g", "float", ""),
            ("C57BL/6 Mouse", "Genotype", "TRUE", "WT", "Genotype of the animal", "", "category", "WT;KO;HET"),
            ("Sprague Dawley Rat", "Tumor Volume", "FALSE", "", "Tumor volume measured weekly", "mm3", "float", "")
        ]
        ws.append([]); ws.append(["Example Data:"])
        for row in example_data:
            ws.append(row)

        for col_idx, col in enumerate(ws.iter_cols()):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx + 1)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(cell.value)
                except: pass
            ws.column_dimensions[column_letter].width = (max_length + 2)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = "animal_model_bulk_upload_template.xlsx"
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return send_file(excel_file, mimetype=mimetype, as_attachment=True, download_name=filename)

    except ImportError:
        flash(_l("The 'openpyxl' library is required. Please install it."), "danger")
        return redirect(url_for('core_models.manage_models'))
    except Exception as e:
        current_app.logger.error(f"Error generating animal model upload template: {e}", exc_info=True)
        flash(_l("An error occurred while generating the template: %(error)s", error=str(e)), "danger")
        return redirect(url_for('core_models.manage_models'))


@core_models_bp.route('/bulk_upload_protocol', methods=['GET', 'POST'])
@login_required
def bulk_upload_protocol():
    if not user_has_permission(current_user, 'CoreModel', 'create', allow_any_team=True):
        flash(_l("You do not have permission to perform this action."), "danger")
        return redirect(url_for('core_models.manage_models'))
    
    form = ProtocolBulkUploadForm()
    if form.validate_on_submit():
        file = form.file.data
        try:
            # Use utility
            data_list, columns = read_excel_to_list(file, sheet_name="Protocol Bulk Upload")
            df = pd.DataFrame(data_list)

            required_cols = ["Protocol Name", "Analyte Name"]
            if not all(col in columns for col in required_cols):
                flash(_l("The uploaded file is missing required columns. Please use the template."), "danger")
                return redirect(url_for('core_models.bulk_upload_protocol'))

            protocols_to_create = {}
            analytes_to_create = {}
            conflicts = []
            protocols_to_skip = []

            for protocol_name, group in df.groupby('Protocol Name'):
                existing_protocol = ProtocolModel.query.filter(func.lower(ProtocolModel.name) == func.lower(protocol_name)).first()
                if existing_protocol:
                    protocols_to_skip.append(protocol_name)
                    continue

                protocol_info = {
                    "name": protocol_name,
                    "severity": group['Severity'].iloc[0],
                    "description": group['Description'].iloc[0],
                    "url": group['URL'].iloc[0],
                    "analytes": []
                }

                if not isinstance(protocol_info['severity'], str) or protocol_info['severity'] not in [s.value for s in Severity]:
                    conflicts.append(f"Protocol '{protocol_name}': Invalid severity value '{protocol_info['severity']}'.")
                
                for index, row in group.iterrows():
                    analyte_name = row['Analyte Name']
                    if not analyte_name:
                        continue

                    analyte_link_info = {
                        "name": analyte_name,
                        "default_value": row.get('Default Value'),
                        "is_metadata": str(row.get('Is Metadata (TRUE/FALSE)')).upper() == 'TRUE'
                    }

                    existing_analyte = Analyte.query.filter(func.lower(Analyte.name) == func.lower(analyte_name)).first()

                    if existing_analyte:
                        if pd.notna(row.get('New Analyte Unit')) and row.get('New Analyte Unit') != existing_analyte.unit:
                            conflicts.append(f"Analyte '{analyte_name}': Unit mismatch. File has '{row.get('New Analyte Unit')}', database has '{existing_analyte.unit}'.")
                        if pd.notna(row.get('New Analyte Data Type')) and row.get('New Analyte Data Type') != existing_analyte.data_type.value:
                            conflicts.append(f"Analyte '{analyte_name}': Data type mismatch. File has '{row.get('New Analyte Data Type')}', database has '{existing_analyte.data_type.value}'.")
                    else:
                        if analyte_name not in analytes_to_create:
                            new_analyte_info = {
                                "name": analyte_name,
                                "description": row.get('New Analyte Description'),
                                "unit": row.get('New Analyte Unit'),
                                "data_type": row.get('New Analyte Data Type'),
                                "allowed_values": row.get('New Analyte Allowed Values')
                            }
                            if not new_analyte_info['data_type'] or new_analyte_info['data_type'] not in [dt.value for dt in AnalyteDataType]:
                                conflicts.append(f"New Analyte '{analyte_name}': Invalid data type '{new_analyte_info['data_type']}'.")
                            if new_analyte_info['data_type'] == 'category' and not new_analyte_info['allowed_values']:
                                conflicts.append(f"New Analyte '{analyte_name}': 'Allowed Values' is required for category data type.")
                            
                            analytes_to_create[analyte_name] = new_analyte_info

                    protocol_info["analytes"].append(analyte_link_info)
                
                protocols_to_create[protocol_name] = protocol_info

            session['bulk_upload_data'] = {
                "protocols": list(protocols_to_create.values()),
                "new_analytes": list(analytes_to_create.values()),
                "conflicts": conflicts,
                "skipped_protocols": protocols_to_skip
            }
            
            return redirect(url_for('core_models.bulk_upload_review'))

        except Exception as e:
            current_app.logger.error(f"Error processing bulk protocol upload: {e}", exc_info=True)
            flash(_l("An error occurred while processing the file: %(error)s", error=str(e)), "danger")
            return redirect(url_for('core_models.bulk_upload_protocol'))

    return render_template('core_models/bulk_upload_protocol.html', form=form)

def generate_protocol_export(protocols):
    all_protocol_data = []
    for protocol in protocols:
        associations = ProtocolAnalyteAssociation.query.filter_by(protocol_model_id=protocol.id).order_by(ProtocolAnalyteAssociation.order).all()

        if not associations:
            all_protocol_data.append({
                "Protocol Name": protocol.name,
                "Severity": protocol.severity.value,
                "Description": protocol.description,
                "URL": protocol.url,
                "Analyte Name": "",
                "Default Value": "",
                "Is Metadata (TRUE/FALSE)": "",
                "New Analyte Description": "",
                "New Analyte Unit": "",
                "New Analyte Data Type": "",
                "New Analyte Allowed Values": ""
            })
        else:
            for assoc in associations:
                analyte = assoc.analyte
                protocol_data = {
                    "Protocol Name": protocol.name,
                    "Severity": protocol.severity.value,
                    "Description": protocol.description,
                    "URL": protocol.url,
                    "Analyte Name": analyte.name,
                    "Default Value": assoc.default_value,
                    "Is Metadata (TRUE/FALSE)": str(assoc.is_metadata).upper(),
                    "New Analyte Description": analyte.description,
                    "New Analyte Unit": analyte.unit,
                    "New Analyte Data Type": analyte.data_type.value,
                    "New Analyte Allowed Values": analyte.allowed_values
                }
                all_protocol_data.append(protocol_data)

    df = pd.DataFrame(all_protocol_data)
    return dataframe_to_excel_bytes(df, sheet_name="Protocol Bulk Upload")

@core_models_bp.route('/download_protocol/<int:protocol_id>')
@login_required
def download_protocol(protocol_id):
    protocol = db.session.get(ProtocolModel, protocol_id)
    if not protocol:
        flash(_l("Protocol not found."), "danger")
        return redirect(url_for('core_models.manage_models'))

    protocols_to_export = [protocol]
    excel_file = generate_protocol_export(protocols_to_export)
    
    filename = f"protocol_{protocol.name}.xlsx"
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return send_file(excel_file, mimetype=mimetype, as_attachment=True, download_name=filename)

@core_models_bp.route('/download_selected_protocols', methods=['POST'])
@login_required
def download_selected_protocols():
    selected_ids = request.form.getlist('selected_protocols')
    if not selected_ids:
        flash(_l("No protocols selected for download."), "warning")
        return redirect(url_for('core_models.manage_models'))

    protocols = ProtocolModel.query.filter(ProtocolModel.id.in_(selected_ids)).all()
    if not protocols:
        flash(_l("Selected protocols not found."), "danger")
        return redirect(url_for('core_models.manage_models'))

    excel_file = generate_protocol_export(protocols)
    
    filename = f"selected_protocols_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return send_file(excel_file, mimetype=mimetype, as_attachment=True, download_name=filename)

@core_models_bp.route('/bulk_upload_review', methods=['GET'])
@login_required
def bulk_upload_review():
    if not user_has_permission(current_user, 'CoreModel', 'create', allow_any_team=True):
        abort(403)
    
    upload_data = session.get('bulk_upload_data', {})
    if not upload_data:
        flash(_l("No upload data found in session. Please start again."), "warning")
        return redirect(url_for('core_models.bulk_upload_protocol'))

    return render_template('core_models/bulk_upload_review.html',
                           protocols=upload_data.get('protocols', []),
                           new_analytes=upload_data.get('new_analytes', []),
                           conflicts=upload_data.get('conflicts', []))

@core_models_bp.route('/bulk_upload_confirm', methods=['POST'])
@login_required
def bulk_upload_confirm():
    if not user_has_permission(current_user, 'CoreModel', 'create', allow_any_team=True):
        abort(403)

    upload_data = session.get('bulk_upload_data', {})
    if not upload_data or upload_data.get('conflicts'):
        flash(_l("Cannot confirm import due to conflicts or missing data."), "danger")
        return redirect(url_for('core_models.bulk_upload_review'))

    try:
        new_analyte_objects = {}
        for analyte_data in upload_data.get('new_analytes', []):
            new_analyte = Analyte(
                name=analyte_data['name'],
                description=analyte_data['description'],
                unit=analyte_data['unit'],
                data_type=[dt for dt in AnalyteDataType if dt.value == analyte_data['data_type']],
                allowed_values=analyte_data['allowed_values'],
                creator_id=current_user.id
            )
            db.session.add(new_analyte)
            new_analyte_objects[analyte_data['name']] = new_analyte
        
        db.session.flush()

        for protocol_data in upload_data.get('protocols', []):
            new_protocol = ProtocolModel(
                name=protocol_data['name'],
                severity=[s for s in Severity if s.value == protocol_data['severity']],
                description=protocol_data['description'],
                url=protocol_data['url']
            )
            db.session.add(new_protocol)
            db.session.flush()

            for analyte_link in protocol_data['analytes']:
                analyte_obj = None
                if analyte_link['name'] in new_analyte_objects:
                    analyte_obj = new_analyte_objects[analyte_link['name']]
                else:
                    analyte_obj = Analyte.query.filter(func.lower(Analyte.name) == func.lower(analyte_link['name'])).first()

                if analyte_obj:
                    association = ProtocolAnalyteAssociation(
                        protocol_model_id=new_protocol.id,
                        analyte_id=analyte_obj.id,
                        default_value=analyte_link['default_value'],
                        is_metadata=analyte_link['is_metadata']
                    )
                    db.session.add(association)

        db.session.commit()
        session.pop('bulk_upload_data', None)
        flash(_l("Bulk import successful!"), "success")
        return redirect(url_for('core_models.manage_models'))

    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error during bulk protocol import confirmation: {e}", exc_info=True)
        flash(_l("An error occurred during the final import: %(error)s", error=str(e)), "danger")
        return redirect(url_for('core_models.bulk_upload_review'))

@core_models_bp.route('/download_protocol_template')
@login_required
def download_protocol_template():
    # Keeping openpyxl logic here for specific formatting
    try:
        from io import BytesIO

        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Protocol Bulk Upload"

        headers = [
            "Protocol Name", "Severity", "Description", "URL",
            "Analyte Name", "Default Value", "Is Metadata (TRUE/FALSE)",
            "New Analyte Description", "New Analyte Unit", "New Analyte Data Type", "New Analyte Allowed Values"
        ]
        ws.append(headers)

        header_font = Font(bold=True, color="FFFFFF")
        protocol_fill = PatternFill(start_color="007bff", end_color="007bff", fill_type="solid")
        analyte_link_fill = PatternFill(start_color="17a2b8", end_color="17a2b8", fill_type="solid")
        new_analyte_fill = PatternFill(start_color="28a745", end_color="28a745", fill_type="solid")

        for i, cell in enumerate(ws[1]):
            cell.font = header_font
            if i < 4:
                cell.fill = protocol_fill
            elif i < 7:
                cell.fill = analyte_link_fill
            else:
                cell.fill = new_analyte_fill

        instructions = [
            "", "Instructions:",
            "1. Each row represents one analyte linked to a protocol.",
            "2. To create one protocol with multiple analytes, repeat the 'Protocol Name' in each row.",
            "3. 'Severity' must be one of: None, Light, Moderate, Severe.",
            "4. 'Is Metadata' should be TRUE or FALSE.",
            "5. The 'New Analyte...' columns should only be filled out if the 'Analyte Name' does not already exist in the system.",
            "6. 'New Analyte Data Type' must be one of: float, int, text, category, date.",
            "7. For 'category' data types, 'New Analyte Allowed Values' must be a semicolon-separated list (e.g., Value1;Value2;Value3)."
        ]
        for instruction in instructions:
            ws.append([instruction])

        example_data = [
            ("Blood Collection Protocol", "Light", "Standard tail vein blood collection.", "http://example.com/blood_sop", "Glucose", "100.5", "FALSE", "", "mg/dL", "float", ""),
            ("Blood Collection Protocol", "Light", "", "", "Lactate", "", "FALSE", "Lactic acid measurement", "mmol/L", "float", ""),
            ("Behavioral Test", "Moderate", "Open field test for anxiety.", "", "Activity Score", "", "TRUE", "Score from 1-10", "", "int", ""),
            ("Behavioral Test", "Moderate", "", "", "Rearing Events", "", "TRUE", "Count of rearing events", "", "int", ""),
            ("New Dosing Protocol", "Light", "Oral gavage of compound X.", "", "Clinical Score", "0", "FALSE", "Daily health score", "", "category", "0;1;2;3;4")
        ]
        
        ws.append([]); ws.append(["Example Data:"])
        for row in example_data:
            ws.append(row)

        for col_idx, col in enumerate(ws.iter_cols()):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx + 1)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(cell.value)
                except: pass
            ws.column_dimensions[column_letter].width = (max_length + 2)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = "protocol_bulk_upload_template.xlsx"
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return send_file(excel_file, mimetype=mimetype, as_attachment=True, download_name=filename)

    except ImportError:
        flash(_l("The 'openpyxl' library is required to generate XLSX files. Please install it."), "danger")
        return redirect(url_for('core_models.manage_models'))
    except Exception as e:
        current_app.logger.error(f"Error generating protocol upload template: {e}", exc_info=True)
        flash(_l("An error occurred while generating the template: %(error)s", error=str(e)), "danger")
        return redirect(url_for('core_models.manage_models'))

@core_models_bp.route('/download_protocol_data/<int:id>')
@login_required
def download_protocol_data(id):
    protocol_model = db.session.get(ProtocolModel, id)
    if not protocol_model:
        flash(_l("Protocol model not found."), "danger")
        return redirect(url_for('core_models.manage_models'))

    data_tables = DataTable.query.filter_by(protocol_id=protocol_model.id).all()

    accessible_data_tables = []
    for dt in data_tables:
        if dt.group and check_group_permission(dt.group, 'read'):
            accessible_data_tables.append(dt)

    if not accessible_data_tables:
        flash(_l("No accessible data tables found for this protocol."), "warning")
        return redirect(url_for('core_models.manage_models'))

    try:
        from io import BytesIO

        import openpyxl

        wb = openpyxl.Workbook()
        if wb.sheetnames == 'Sheet': 
            wb.remove(wb.active)
        
        ws = wb.create_sheet(title=f"{protocol_model.name[:25]}_Data")

        all_columns_animal = set()
        
        ordered_protocol_cols = []
        if protocol_model.analyte_associations:
            associations = ProtocolAnalyteAssociation.query.filter_by(protocol_model_id=protocol_model.id).order_by(ProtocolAnalyteAssociation.order).all()
            ordered_protocol_cols.extend(assoc.analyte.name for assoc in associations)

        ordered_animal_cols = []
        if 'ID' in all_columns_animal:
            ordered_animal_cols.append('ID')
            all_columns_animal.discard('ID') 
        ordered_animal_cols.extend(sorted(list(all_columns_animal)))
        
        header_row = ["Experimental Group", "DataTable Date"] + ordered_animal_cols + ordered_protocol_cols
        ws.append(header_row)

        for data_table in accessible_data_tables:
            group = data_table.group
            group_name = group.name if group else "N/A"
            dt_date_str = data_table.date.strftime('%Y-%m-%d') if hasattr(data_table.date, 'strftime') else str(data_table.date)
            
            group_animal_data_list = group.animal_data if group else []
            experiment_rows_for_dt = {row.row_index: row.row_data for row in data_table.experiment_rows.all()}

            for i in range(len(group_animal_data_list)):
                animal_data_dict = group_animal_data_list[i] if i < len(group_animal_data_list) else {}
                protocol_row_data_dict = experiment_rows_for_dt.get(i, {})
                
                excel_row = [group_name, dt_date_str]
                for col_name in ordered_animal_cols:
                    excel_row.append(animal_data_dict.get(col_name, "")) 
                for col_name in ordered_protocol_cols:
                    excel_row.append(protocol_row_data_dict.get(col_name, "")) 
                
                ws.append(excel_row)

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        filename = f"{protocol_model.name}_aggregated_data.xlsx".replace(" ", "_")
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return send_file(excel_file, mimetype=mimetype, as_attachment=True, download_name=filename)

    except ImportError:
        flash(_l("Please install openpyxl and pandas libraries to generate XLSX files."), "danger")
        return redirect(url_for('core_models.manage_models'))
    except Exception as e:
        current_app.logger.error(f"Error generating XLSX file for protocol {id}: {e}", exc_info=True)
        flash(_l("Error generating XLSX file: {}").format(str(e)), "danger")
        return redirect(url_for('core_models.manage_models'))

@core_models_bp.route('/', methods=['GET'])
@login_required
def manage_models():
    can_manage = user_has_permission(current_user, 'CoreModel', 'edit', allow_any_team=True) or user_has_permission(current_user, 'CoreModel', 'delete', allow_any_team=True)
    
    if user_has_permission(current_user, 'CoreModel', 'read', allow_any_team=True):
        animal_models_query = AnimalModel.query.order_by(AnimalModel.name).all()
        protocol_models_query = ProtocolModel.query.order_by(ProtocolModel.name).all()
    else:
        accessible_projects = current_user.get_accessible_projects(include_archived=True)
        accessible_project_ids = [p.id for p in accessible_projects]

        if accessible_project_ids:
            animal_model_ids = db.session.query(ExperimentalGroup.model_id).filter(ExperimentalGroup.project_id.in_(accessible_project_ids)).distinct().all()
            animal_model_ids = [id[0] for id in animal_model_ids]
            animal_models_query = AnimalModel.query.filter(AnimalModel.id.in_(animal_model_ids)).order_by(AnimalModel.name).all()

            protocol_model_ids = db.session.query(DataTable.protocol_id).join(ExperimentalGroup).filter(ExperimentalGroup.project_id.in_(accessible_project_ids)).distinct().all()
            protocol_model_ids = [id[0] for id in protocol_model_ids]
            protocol_models_query = ProtocolModel.query.filter(ProtocolModel.id.in_(protocol_model_ids)).order_by(ProtocolModel.name).all()
        else:
            animal_models_query = []
            protocol_models_query = []

    animal_models_data = []
    for model_item in animal_models_query:
        animal_models_data.append({
            'id': model_item.id,
            'name': model_item.name,
            'related_groups_count': model_item.groups.count() 
        })

    protocol_models_data = []
    for model_item in protocol_models_query:
        protocol_models_data.append({
            'id': model_item.id,
            'name': model_item.name,
            'severity': model_item.severity.value if model_item.severity else _l('N/A'), 
            'related_datatables_count': model_item.data_tables.count() 
        })

    return render_template('core_models/manage_models.html',
                           animal_models_data=animal_models_data, 
                           protocol_models_data=protocol_models_data,
                           can_manage=can_manage)


@core_models_bp.route('/edit/<string:model_type>', methods=['GET', 'POST'])
@core_models_bp.route('/edit/<string:model_type>/<int:id>', methods=['GET', 'POST'])
@login_required
def edit_model(model_type, id=None):
    required_permission = 'edit' if id else 'create'
    if not user_has_permission(current_user, 'CoreModel', required_permission, allow_any_team=True):
        flash(_l("You do not have permission to perform this action."), "danger")
        return redirect(url_for('core_models.manage_models'))
    
    ModelClass = AnimalModel if model_type == 'animal' else ProtocolModel
    model = db.session.get(ModelClass, id) if id else None
    FormClass = ProtocolModelForm if model_type == 'protocol' else CoreModelForm
    
    form = FormClass(obj=model)
    
    if request.method == 'GET' and model and model_type == 'protocol':
        form.severity.data = model.severity.name
        form.description.data = model.description
        form.url.data = model.url
        if model.external_skill_ids:
             # Ensure we pass a list of ints
             form.external_skills.data = model.external_skill_ids
             
        # Pre-fill controlled molecules
        if model.molecule_associations:
            form.controlled_molecules.data = [assoc.molecule_id for assoc in model.molecule_associations]
        
        if model.import_pipelines:
            form.import_pipelines.data = [p.id for p in model.import_pipelines]

    if form.validate_on_submit():
        name_from_form = form.name.data.strip()
        unique_name = generate_unique_name(name_from_form, ModelClass.query.filter(ModelClass.id != id) if id else ModelClass.query)
        if unique_name != name_from_form:
            flash(_l(f"Model name '{name_from_form}' was already taken. Renamed to '{unique_name}'."), "warning")
            name_from_form = unique_name

        if not model:
            model = ModelClass(name=name_from_form)
            if model_type == 'protocol':
                model.severity = Severity[form.severity.data]
                model.description = form.description.data
                model.url = form.url.data
                model.enable_import_wizard = form.enable_import_wizard.data
            db.session.add(model)
            try:
                db.session.flush()
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error flushing session for new model: {e}", exc_info=True)
                flash(_l('A database error occurred while creating the model. The model was not saved.'), 'danger')
                all_analytes = Analyte.query.order_by(Analyte.name).all()
                all_analytes_data = [{'id': a.id, 'name': a.name, 'description': a.description, 'unit': a.unit, 'data_type': a.data_type.name, 'allowed_values': a.allowed_values} for a in all_analytes]
                return render_template('core_models/edit_model.html', model=model, model_type=model_type, form=form, data_types=list(AnalyteDataType), all_analytes_data=all_analytes_data)
        else:
            model.name = name_from_form
            if model_type == 'protocol':
                model.severity = Severity[form.severity.data]
                model.description = form.description.data
                model.url = form.url.data
                model.enable_import_wizard = form.enable_import_wizard.data

        if model_type == 'protocol':
            attachment_file = request.files.get('attachment')
            if attachment_file and attachment_file.filename != '':
                try:
                    # Use utility for validation
                    validate_file_type(attachment_file)
                    
                    for old_attachment in model.attachments:
                        try:
                            if os.path.exists(old_attachment.filepath):
                                os.remove(old_attachment.filepath)
                        except OSError as e:
                            current_app.logger.error(f"Error deleting old attachment file {old_attachment.filepath}: {e}")
                        db.session.delete(old_attachment)

                    filename = secure_filename(attachment_file.filename)
                    upload_folder = os.path.join(current_app.config['UPLOAD_FOLDER'], 'protocols')
                    os.makedirs(upload_folder, exist_ok=True)
                    
                    unique_filename = f"{model.id}_{filename}"
                    full_filepath = os.path.join(upload_folder, unique_filename)
                    relative_filepath = os.path.join('protocols', unique_filename)
                    
                    attachment_file.save(full_filepath)
                    
                    new_attachment = ProtocolAttachment(
                        protocol_id=model.id,
                        filename=filename,
                        filepath=relative_filepath
                    )
                    db.session.add(new_attachment)
                except ValueError as e:
                    flash(_l(e), 'danger')

        # Process analytes
        submitted_analyte_ids_str = request.form.getlist('analytes')
        final_analyte_ids = []
        new_analyte_map = {} 

        for temp_id in submitted_analyte_ids_str:
            if temp_id.startswith('new_'):
                details_json = request.form.get(f'new_analyte_details_{temp_id}')
                if details_json:
                    details = json.loads(details_json)
                    analyte_name = generate_unique_name(details['name'], Analyte.query)
                    if analyte_name != details['name']:
                        flash(_l(f"Analyte name '{details['name']}' was already taken. Renamed to '{analyte_name}'."), "warning")
                    
                    new_analyte = Analyte(
                        name=analyte_name,
                        description=details.get('description'),
                        unit=details.get('unit'),
                        data_type=AnalyteDataType[details['data_type']],
                        allowed_values=details.get('allowed_values'),
                        creator_id=current_user.id
                    )
                    db.session.add(new_analyte)
                    new_analyte_map[temp_id] = new_analyte
            else:
                try:
                    final_analyte_ids.append(int(temp_id))
                except ValueError:
                    continue
        
        if new_analyte_map:
            try:
                db.session.flush()
            except Exception as e:
                current_app.logger.error(f"Error during session flush when creating analytes: {e}", exc_info=True)
                db.session.rollback()
                flash(_l('A database error occurred while creating a new analyte. The model was not saved.'), 'danger')
                return render_template('core_models/edit_model.html', model=model, model_type=model_type, form=form, data_types=list(AnalyteDataType))

            for temp_id, analyte_obj in new_analyte_map.items():
                final_analyte_ids.append(analyte_obj.id)
                new_analyte_map[temp_id] = analyte_obj.id


        if model_type == 'protocol':
            existing_associations = {assoc.analyte_id: assoc for assoc in model.analyte_associations}
            submitted_ids_set = set(final_analyte_ids)

            for analyte_id, association in existing_associations.items():
                if analyte_id not in submitted_ids_set:
                    db.session.delete(association)

            for i, analyte_id in enumerate(final_analyte_ids):
                original_id_str = submitted_analyte_ids_str[i] 
                
                default_value = request.form.get(f'default_value_{original_id_str}')
                calculation_formula = request.form.get(f'calculation_formula_{original_id_str}')
                is_metadata = f'is_metadata_{original_id_str}' in request.form
                is_sensitive = f'is_sensitive_{original_id_str}' in request.form
                
                analyte = db.session.get(Analyte, analyte_id)
                if not analyte:
                    continue

                if analyte.data_type == AnalyteDataType.CATEGORY and analyte.allowed_values:
                    allowed = [v.strip() for v in analyte.allowed_values.split(';')]
                    if default_value and default_value not in allowed:
                        flash(_l(f"Invalid default value '{default_value}' for analyte '{analyte.name}'. It has been reset."), "warning")
                        default_value = ""
                
                if default_value:
                    if analyte.data_type == AnalyteDataType.FLOAT:
                        try:
                            float(default_value)
                        except ValueError:
                            flash(_l(f"Invalid float value '{default_value}' for analyte '{analyte.name}'. It has been reset."), "warning")
                            default_value = ""
                    elif analyte.data_type == AnalyteDataType.INT:
                        try:
                            int(default_value)
                        except ValueError:
                            flash(_l(f"Invalid integer value '{default_value}' for analyte '{analyte.name}'. It has been reset."), "warning")
                            default_value = ""
                
                analyte.is_sensitive = is_sensitive

                association = existing_associations.get(analyte_id)
                if association:
                    association.default_value = default_value
                    association.calculation_formula = calculation_formula
                    association.is_metadata = is_metadata
                    association.order = i
                else:
                    new_assoc = ProtocolAnalyteAssociation(
                        protocol_model_id=model.id,
                        analyte_id=analyte_id,
                        default_value=default_value,
                        calculation_formula=calculation_formula,
                        is_metadata=is_metadata,
                        order=i
                    )
                    db.session.add(new_assoc)

            # Handle required skills for protocols
            if model_type == 'protocol':
                required_skills_list = request.form.getlist('required_skills')
                if required_skills_list:
                    skill_ids = []
                    for skill_id_str in required_skills_list:
                        try:
                            skill_ids.append(int(skill_id_str.strip()))
                        except ValueError:
                            continue
                    model.external_skill_ids = skill_ids
                    flag_modified(model, 'external_skill_ids')
                else:
                    model.external_skill_ids = None
                    flag_modified(model, 'external_skill_ids')
                    model.external_skill_ids = None
                    flag_modified(model, 'external_skill_ids')

            # Handle controlled molecules associations
            if model_type == 'protocol':
                submitted_molecule_ids = form.controlled_molecules.data or []
                existing_mol_assocs = {assoc.molecule_id: assoc for assoc in model.molecule_associations}
                
                # Add new associations
                for mol_id in submitted_molecule_ids:
                    if mol_id not in existing_mol_assocs:
                        new_assoc = ProtocolMoleculeAssociation(protocol_id=model.id, molecule_id=mol_id)
                        db.session.add(new_assoc)
                
                # Remove unselected associations
                for mol_id, assoc in existing_mol_assocs.items():
                    if mol_id not in submitted_molecule_ids:
                        db.session.delete(assoc)

            # Handle import pipelines associations
            if model_type == 'protocol':
                submitted_pipeline_ids = form.import_pipelines.data or []
                model.import_pipelines = ImportPipeline.query.filter(ImportPipeline.id.in_(submitted_pipeline_ids)).all()

            try:
                db.session.commit()
                flash(_l('Protocol Model "%(name)s" updated successfully!', name=model.name), 'success')
                return redirect(url_for('core_models.edit_model', model_type='protocol', id=model.id))
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error saving protocol model {model.id} analyte order: {e}", exc_info=True)
                flash(_l('Error saving protocol model analytes: %(error)s', error=str(e)), 'danger')
        else: 
            AnimalModelAnalyteAssociation.query.filter_by(animal_model_id=model.id).delete()
            db.session.flush()

            for i, analyte_id in enumerate(final_analyte_ids):
                original_id_str = submitted_analyte_ids_str[i]
                is_metadata = f'is_metadata_{original_id_str}' in request.form
                default_value = request.form.get(f'default_value_{original_id_str}')

                analyte = db.session.get(Analyte, analyte_id)
                if not analyte:
                    continue

                if analyte.data_type == AnalyteDataType.CATEGORY and analyte.allowed_values:
                    allowed = [v.strip() for v in analyte.allowed_values.split(';')]
                    if default_value and default_value not in allowed:
                        flash(_l(f"Invalid default value '{default_value}' for analyte '{analyte.name}'. It has been reset."), "warning")
                        default_value = ""
                
                if default_value:
                    if analyte.data_type == AnalyteDataType.FLOAT:
                        try:
                            float(default_value)
                        except ValueError:
                            flash(_l(f"Invalid float value '{default_value}' for analyte '{analyte.name}'. It has been reset."), "warning")
                            default_value = ""
                    elif analyte.data_type == AnalyteDataType.INT:
                        try:
                            int(default_value)
                        except ValueError:
                            flash(_l(f"Invalid integer value '{default_value}' for analyte '{analyte.name}'. It has been reset."), "warning")
                            default_value = ""

                analyte.is_metadata = is_metadata
                analyte.default_value = default_value
                is_sensitive = f'is_sensitive_{original_id_str}' in request.form
                analyte.is_sensitive = is_sensitive

                new_assoc = AnimalModelAnalyteAssociation(
                    animal_model_id=model.id,
                    analyte_id=analyte_id,
                    order=i
                )
                db.session.add(new_assoc)

            mandatory_analytes = Analyte.query.filter_by(is_mandatory=True).all()
            current_max_order = len(final_analyte_ids)
            for man_analyte in mandatory_analytes:
                if man_analyte.id not in final_analyte_ids:
                    new_mandatory_assoc = AnimalModelAnalyteAssociation(
                        animal_model_id=model.id,
                        analyte_id=man_analyte.id,
                        order=current_max_order
                    )
                    db.session.add(new_mandatory_assoc)
                    current_max_order += 1

            try:
                db.session.commit()
                flash(_l('Animal Model "%(name)s" updated successfully!', name=model.name), 'success')
                return redirect(url_for('core_models.edit_model', model_type='animal', id=model.id))
            except Exception as e:
                db.session.rollback()
                current_app.logger.error(f"Error saving animal model {model.id} analyte order and details: {e}", exc_info=True)
                flash(_l('Error saving animal model analytes: %(error)s', error=str(e)), 'danger')

    all_analytes = Analyte.query.order_by(Analyte.name).all()
    all_analytes_data = [{'id': a.id, 'name': a.name, 'description': a.description, 'unit': a.unit, 'data_type': a.data_type.name, 'allowed_values': a.allowed_values} for a in all_analytes]

    selected_analytes_data = []
    if model:
        if model_type == 'animal':
            associations = AnimalModelAnalyteAssociation.query.filter_by(animal_model_id=model.id).order_by(AnimalModelAnalyteAssociation.order).all()
            for assoc in associations:
                selected_analytes_data.append({
                    'id': assoc.analyte.id,
                    'name': assoc.analyte.name,
                    'description': assoc.analyte.description,
                    'unit': assoc.analyte.unit,
                    'data_type': assoc.analyte.data_type.name,
                    'allowed_values': assoc.analyte.allowed_values,
                    'default_value': assoc.analyte.default_value,
                    'is_metadata': assoc.analyte.is_metadata,
                    'is_sensitive': assoc.analyte.is_sensitive
                })
        elif model_type == 'protocol':
            associations = ProtocolAnalyteAssociation.query.filter_by(protocol_model_id=model.id).order_by(ProtocolAnalyteAssociation.order).all()
            for assoc in associations:
                selected_analytes_data.append({
                    'id': assoc.analyte.id,
                    'name': assoc.analyte.name,
                    'description': assoc.analyte.description,
                    'unit': assoc.analyte.unit,
                    'data_type': assoc.analyte.data_type.name,
                    'allowed_values': assoc.analyte.allowed_values,
                    'default_value': assoc.default_value,
                    'calculation_formula': assoc.calculation_formula,
                    'is_metadata': assoc.is_metadata,
                    'is_sensitive': assoc.analyte.is_sensitive
                })

    return render_template('core_models/edit_model.html', model=model, model_type=model_type, form=form, data_types=list(AnalyteDataType), all_analytes_data=all_analytes_data, selected_analytes_data=selected_analytes_data)


@core_models_bp.route('/protocols/<int:protocol_id>/attachments/<int:attachment_id>')
@login_required
def download_protocol_attachment(protocol_id, attachment_id):
    protocol = db.session.get(ProtocolModel, protocol_id)
    if not protocol:
        abort(404)
    if not user_has_permission(current_user, 'CoreModel', 'edit', allow_any_team=True):
        flash(_l("You do not have permission to view this attachment."), "danger")
        abort(403)

    attachment = db.session.get(ProtocolAttachment, attachment_id)
    if not attachment or attachment.protocol_id != protocol.id:
        abort(404)

    upload_folder = current_app.config.get('UPLOAD_FOLDER')
    if not upload_folder:
        current_app.logger.error("UPLOAD_FOLDER is not configured.")
        abort(500)

    try:
        return send_from_directory(
            upload_folder,
            attachment.filepath,
            as_attachment=True,
            download_name=attachment.filename
        )
    except FileNotFoundError:
        current_app.logger.error(f"Attachment file not found at path: {os.path.join(upload_folder, attachment.filepath)}")
        abort(404)


@core_models_bp.route('/delete/<string:model_type>/<int:id>', methods=['POST'])
@login_required
def delete_model(model_type, id):
    if not user_has_permission(current_user, 'CoreModel', 'delete', allow_any_team=True):
        flash(_l("You do not have permission to delete Core Models."), "danger")
        return redirect(url_for('core_models.manage_models'))
    ModelClass = AnimalModel if model_type == 'animal' else ProtocolModel if model_type == 'protocol' else None
    if not ModelClass:
        flash(_l("Invalid model type specified."), "danger"); return redirect(url_for('core_models.manage_models'))

    model = db.session.get(ModelClass, id)
    if not model:
        flash(_l(f"{model_type.capitalize()} model not found."), "danger"); return redirect(url_for('core_models.manage_models'))

    dependency_exists = False
    if model_type == 'animal':
        dependency_exists = model.groups.first() is not None
        dependency_message = _l('Cannot delete model because it is in use by experimental groups.')
    elif model_type == 'protocol':
        dependency_exists = model.data_tables.first() is not None
        dependency_message = _l('Cannot delete model because it is in use by data tables.')

    if dependency_exists:
        flash(dependency_message, 'danger'); return redirect(url_for('core_models.manage_models'))

    try:
        db.session.delete(model)
        db.session.commit()
        flash(_l(f'{model_type.capitalize()} model deleted successfully.'), 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting {model_type} model ID {id}: {e}", exc_info=True)
        flash(_l(f'Error deleting model: {e}'), 'danger')
    return redirect(url_for('core_models.manage_models'))


@core_models_bp.route('/duplicate/<string:model_type>/<int:id>', methods=['POST'])
@login_required
def duplicate_model(model_type, id):
    if not user_has_permission(current_user, 'CoreModel', 'create', allow_any_team=True):
        flash(_l("You do not have permission to duplicate Core Models."), "danger")
        return redirect(url_for('core_models.manage_models'))
    try:
        validate_csrf(request.form.get('csrf_token'))
    except Exception as e:
        current_app.logger.error(f"CSRF validation failed during model duplication: {e}")
        flash(_l('CSRF token is missing or invalid. Could not duplicate model.'), 'danger')
        return redirect(url_for('core_models.manage_models'))

    new_name = request.form.get('new_model_name', '').strip()
    if not new_name:
        flash(_l("New model name cannot be empty."), "danger")
        return redirect(url_for('core_models.manage_models'))

    ModelClass = AnimalModel if model_type == 'animal' else ProtocolModel if model_type == 'protocol' else None
    if not ModelClass:
        flash(_l("Invalid model type specified for duplication."), "danger")
        return redirect(url_for('core_models.manage_models'))

    original_model = db.session.get(ModelClass, id)
    if not original_model:
        flash(_l(f"Original {model_type.capitalize()} model not found for duplication."), "danger")
        return redirect(url_for('core_models.manage_models'))

    other_models_query = ModelClass.query
    unique_new_name = generate_unique_name(new_name, other_models_query)
    if unique_new_name != new_name:
        flash(_l(f"The desired name '{new_name}' was already taken. The model has been duplicated as '{unique_new_name}' instead."), "warning")
        new_name = unique_new_name

    new_model = ModelClass(name=new_name)
    if model_type == 'protocol':
        if hasattr(original_model, 'severity'):
            new_model.severity = original_model.severity
        new_model.enable_import_wizard = original_model.enable_import_wizard
        for assoc in original_model.analyte_associations:
            new_assoc = ProtocolAnalyteAssociation(
                protocol_model=new_model,
                analyte=assoc.analyte,
                default_value=assoc.default_value,
                calculation_formula=assoc.calculation_formula,
                is_metadata=assoc.is_metadata,
                order=assoc.order
            )
            db.session.add(new_assoc)
    else: 
        associations = AnimalModelAnalyteAssociation.query.filter_by(animal_model_id=original_model.id).order_by(AnimalModelAnalyteAssociation.order).all()
        for assoc in associations:
            new_assoc = AnimalModelAnalyteAssociation(
                animal_model=new_model,
                analyte=assoc.analyte,
                order=assoc.order
            )
            db.session.add(new_assoc)

    if model_type == 'protocol' and hasattr(original_model, 'severity'):
        new_model.severity = original_model.severity

    try:
        db.session.add(new_model)
        db.session.commit()
        flash(_l(f'{model_type.capitalize()} model \'{original_model.name}\' duplicated successfully as \'{new_model.name}\'.'), 'success')
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error duplicating {model_type} model '{original_model.name}' to '{new_model.name}': {e}", exc_info=True)
        flash(_l(f'Error duplicating model: {str(e)}'), 'danger')
    return redirect(url_for('core_models.manage_models'))
    
@core_models_bp.route('/download_template/<int:model_id>')
@login_required
def download_group_template(model_id):
    model = db.session.get(AnimalModel, model_id)
    if not model:
        flash(_l("Animal model not found."), "danger"); return redirect(url_for('core_models.manage_models'))

    analytes = model.analytes
    if not analytes:
         flash(_l("Cannot generate template: Animal model has no analytes defined."), "warning")
         return redirect(url_for('core_models.edit_model', model_type='animal', id=model_id))

    try:
        excel_file = generate_xlsx_template(analytes)
        filename = f'{model.name}_group_upload_template.xlsx'.replace(' ', '_')
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return send_file(excel_file, mimetype=mimetype, as_attachment=True, download_name=filename)
    except Exception as e:
        current_app.logger.error(f"Error generating template for model {model_id}: {e}", exc_info=True)
        flash(_l("Error generating template file."), "danger")
        return redirect(url_for('core_models.manage_models'))

@core_models_bp.route('/api/proxy/tm_skills')
@login_required
def proxy_tm_skills():
    connector = TrainingManagerConnector()
    skills = connector.get_skills()
    if skills:
        return skills
    else:
        return {'error': 'Unable to fetch skills from Training Manager'}, 503
