# app/controlled_molecules/routes.py
"""Routes for controlled molecules management."""
from io import BytesIO

from flask import (current_app, flash, make_response, redirect,
                   render_template, request, send_file, url_for)
from flask_babel import lazy_gettext as _l
from flask_login import current_user, login_required
from sqlalchemy import and_, or_

from app.decorators import permission_required
from app.extensions import db
from app.forms.controlled_molecules import ControlledMoleculeForm
from app.models import (AuditLog, ControlledMolecule, DataTable,
                         DataTableMoleculeUsage, ExperimentalGroup,
                         ProtocolMoleculeAssociation, RegulationCategory,
                         user_has_permission)
from app.services.audit_service import log_action

from . import controlled_molecules_bp


@controlled_molecules_bp.route('/')
@login_required
@permission_required('ControlledMolecule', 'View')
def list_molecules():
    """List all controlled molecules with filtering."""
    # Get filter parameters
    category_filter = request.args.get('category', '')
    active_filter = request.args.get('active', 'all')
    search_query = request.args.get('search', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    # Build query
    query = ControlledMolecule.query
    
    if category_filter:
        query = query.filter(ControlledMolecule.regulation_category == RegulationCategory[category_filter])
    
    if active_filter == 'active':
        query = query.filter(ControlledMolecule.is_active == True)
    elif active_filter == 'inactive':
        query = query.filter(ControlledMolecule.is_active == False)
    
    if search_query:
        query = query.filter(or_(
            ControlledMolecule.name.ilike(f'%{search_query}%'),
            ControlledMolecule.internal_reference.ilike(f'%{search_query}%'),
            ControlledMolecule.cas_number.ilike(f'%{search_query}%')
        ))

    # Filter by usage dates if provided
    if start_date or end_date:
        usage_subquery = db.session.query(DataTableMoleculeUsage.molecule_id).join(
            DataTable, DataTableMoleculeUsage.data_table_id == DataTable.id
        )
        if start_date:
            usage_subquery = usage_subquery.filter(DataTable.date >= start_date)
        if end_date:
            usage_subquery = usage_subquery.filter(DataTable.date <= end_date)
        query = query.filter(ControlledMolecule.id.in_(usage_subquery.subquery()))

    molecules = query.order_by(ControlledMolecule.name).all()
    
    # Check permissions for create/edit/delete
    can_create = user_has_permission(current_user, 'ControlledMolecule', 'Create')
    can_edit = user_has_permission(current_user, 'ControlledMolecule', 'Edit')
    can_delete = user_has_permission(current_user, 'ControlledMolecule', 'Delete')
    
    return render_template('controlled_molecules/list_molecules.html',
                          molecules=molecules,
                          categories=RegulationCategory,
                          can_create=can_create,
                          can_edit=can_edit,
                          can_delete=can_delete,
                          category_filter=category_filter,
                          active_filter=active_filter,
                          search_query=search_query)


@controlled_molecules_bp.route('/export')
@login_required
@permission_required('ControlledMolecule', 'View')
def export_molecules():
    """Export controlled molecules to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    # Get same filters as list_molecules
    category_filter = request.args.get('category', '')
    active_filter = request.args.get('active', 'all')
    search_query = request.args.get('search', '')
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    query = ControlledMolecule.query
    
    if category_filter:
        query = query.filter(ControlledMolecule.regulation_category == RegulationCategory[category_filter])
    
    if active_filter == 'active':
        query = query.filter(ControlledMolecule.is_active == True)
    elif active_filter == 'inactive':
        query = query.filter(ControlledMolecule.is_active == False)
    
    if search_query:
        query = query.filter(or_(
            ControlledMolecule.name.ilike(f'%{search_query}%'),
            ControlledMolecule.internal_reference.ilike(f'%{search_query}%'),
            ControlledMolecule.cas_number.ilike(f'%{search_query}%')
        ))
    
    if start_date or end_date:
        usage_subquery = db.session.query(DataTableMoleculeUsage.molecule_id).join(
            DataTable, DataTableMoleculeUsage.data_table_id == DataTable.id
        )
        if start_date:
            usage_subquery = usage_subquery.filter(DataTable.date >= start_date)
        if end_date:
            usage_subquery = usage_subquery.filter(DataTable.date <= end_date)
        query = query.filter(ControlledMolecule.id.in_(usage_subquery.subquery()))
    
    molecules = query.order_by(ControlledMolecule.name).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Controlled Molecules"
    
    # Headers
    headers = [
        'Name', 'Category', 'Internal Ref', 'CAS Number', 'Unit', 'Storage Location', 'Responsible', 'Status', 'Created At', 'Updated At'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for row_num, molecule in enumerate(molecules, 2):
        ws.cell(row=row_num, column=1, value=molecule.name)
        ws.cell(row=row_num, column=2, value=molecule.regulation_category.value)
        ws.cell(row=row_num, column=3, value=molecule.internal_reference or '')
        ws.cell(row=row_num, column=4, value=molecule.cas_number or '')
        ws.cell(row=row_num, column=5, value=molecule.unit or '')
        ws.cell(row=row_num, column=6, value=molecule.storage_location or '')
        ws.cell(row=row_num, column=7, value=molecule.responsible.email if molecule.responsible else '')
        ws.cell(row=row_num, column=8, value='Active' if molecule.is_active else 'Inactive')
        ws.cell(row=row_num, column=9, value=molecule.created_at.strftime('%Y-%m-%d %H:%M') if molecule.created_at else '')
        ws.cell(row=row_num, column=10, value=molecule.updated_at.strftime('%Y-%m-%d %H:%M') if molecule.updated_at else '')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"controlled_molecules_{category_filter or 'all'}_{active_filter}_{start_date or 'all'}_{end_date or 'all'}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@controlled_molecules_bp.route('/create', methods=['GET', 'POST'])
@login_required
@permission_required('ControlledMolecule', 'Create')
def create_molecule():
    """Create a new controlled molecule."""
    form = ControlledMoleculeForm()
    
    if form.validate_on_submit():
        try:
            molecule = ControlledMolecule(
                name=form.name.data,
                regulation_category=RegulationCategory[form.regulation_category.data],
                storage_location=form.storage_location.data,
                responsible_id=form.responsible_id.data if form.responsible_id.data else None,
                cas_number=form.cas_number.data,
                internal_reference=form.internal_reference.data,
                unit=form.unit.data,
                requires_secure_prescription=form.requires_secure_prescription.data,
                max_prescription_days=form.max_prescription_days.data,
                notes=form.notes.data,
                is_active=form.is_active.data
            )
            
            db.session.add(molecule)
            db.session.commit()
            
            log_action(
                resource_type='ControlledMolecule',
                resource_id=molecule.id,
                action='create',
                details=f"Created controlled molecule: {molecule.name}"
            )
            
            flash(_l('Controlled molecule created successfully.'), 'success')
            return redirect(url_for('controlled_molecules.list_molecules'))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error creating controlled molecule: {e}")
            flash(_l('Error creating controlled molecule.'), 'danger')
    
    return render_template('controlled_molecules/create_molecule.html', form=form)


@controlled_molecules_bp.route('/<int:id>/edit', methods=['GET', 'POST'])
@login_required
@permission_required('ControlledMolecule', 'Edit')
def edit_molecule(id):
    """Edit an existing controlled molecule."""
    molecule = ControlledMolecule.query.get_or_404(id)
    form = ControlledMoleculeForm(obj=molecule)
    
    if form.validate_on_submit():
        try:
            old_data = molecule.to_dict()
            
            molecule.name = form.name.data
            molecule.regulation_category = RegulationCategory[form.regulation_category.data]
            molecule.storage_location = form.storage_location.data
            molecule.responsible_id = form.responsible_id.data if form.responsible_id.data else None
            molecule.cas_number = form.cas_number.data
            molecule.internal_reference = form.internal_reference.data
            molecule.unit = form.unit.data
            molecule.requires_secure_prescription = form.requires_secure_prescription.data
            molecule.max_prescription_days = form.max_prescription_days.data
            molecule.notes = form.notes.data
            molecule.is_active = form.is_active.data
            
            db.session.commit()
            
            log_action(
                resource_type='ControlledMolecule',
                resource_id=molecule.id,
                action='update',
                details=f"Updated controlled molecule: {molecule.name}",
                old_value=old_data,
                new_value=molecule.to_dict()
            )
            
            flash(_l('Controlled molecule updated successfully.'), 'success')
            return redirect(url_for('controlled_molecules.list_molecules'))
            
        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error updating controlled molecule: {e}")
            flash(_l('Error updating controlled molecule.'), 'danger')
    
    elif request.method == 'GET':
        # Pre-populate form
        form.regulation_category.data = molecule.regulation_category.name
        form.responsible_id.data = molecule.responsible_id
    
    return render_template('controlled_molecules/edit_molecule.html', form=form, molecule=molecule)


@controlled_molecules_bp.route('/<int:id>/delete', methods=['POST'])
@login_required
@permission_required('ControlledMolecule', 'Delete')
def delete_molecule(id):
    """Delete a controlled molecule."""
    molecule = ControlledMolecule.query.get_or_404(id)
    
    # Check if molecule has usage records
    usage_count = DataTableMoleculeUsage.query.filter_by(molecule_id=id).count()
    
    if usage_count > 0:
        flash(_l('Cannot delete molecule: it has %(count)d usage record(s). Consider deactivating instead.', count=usage_count), 'warning')
        return redirect(url_for('controlled_molecules.list_molecules'))
    
    try:
        molecule_name = molecule.name
        db.session.delete(molecule)
        db.session.commit()
        
        log_action(
            resource_type='ControlledMolecule',
            resource_id=id,
            action='delete',
            details=f"Deleted controlled molecule: {molecule_name}"
        )
        
        flash(_l('Controlled molecule deleted successfully.'), 'success')
        
    except Exception as e:
        db.session.rollback()
        current_app.logger.error(f"Error deleting controlled molecule: {e}")
        flash(_l('Error deleting controlled molecule.'), 'danger')
    
    return redirect(url_for('controlled_molecules.list_molecules'))


@controlled_molecules_bp.route('/<int:id>/usage_history')
@login_required
@permission_required('ControlledMolecule', 'View')
def usage_history(id):
    """View usage history for a specific molecule."""
    molecule = ControlledMolecule.query.get_or_404(id)

    # Get filter parameters
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')

    # Get all usage records with related data
    query = db.session.query(DataTableMoleculeUsage, DataTable, ExperimentalGroup).join(
        DataTable, DataTableMoleculeUsage.data_table_id == DataTable.id
    ).join(
        ExperimentalGroup, DataTable.group_id == ExperimentalGroup.id
    ).filter(
        DataTableMoleculeUsage.molecule_id == id
    )

    # Apply date filters
    if start_date:
        query = query.filter(DataTable.date >= start_date)
    if end_date:
        query = query.filter(DataTable.date <= end_date)

    usages = query.order_by(DataTableMoleculeUsage.recorded_at.desc()).all()
    
    # Calculate totals
    total_volume = sum(usage[0].volume_used for usage in usages if usage[0].volume_used)
    total_animals = sum(usage[0].number_of_animals for usage in usages)
    
    return render_template('controlled_molecules/usage_history.html',
                          molecule=molecule,
                          usages=usages,
                          total_volume=total_volume,
                          total_animals=total_animals)


@controlled_molecules_bp.route('/<int:id>/export_usage_history')
@login_required
@permission_required('ControlledMolecule', 'View')
def export_usage_history(id):
    """Export usage history for a specific molecule to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    molecule = ControlledMolecule.query.get_or_404(id)
    
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    
    query = db.session.query(DataTableMoleculeUsage, DataTable, ExperimentalGroup).join(
        DataTable, DataTableMoleculeUsage.data_table_id == DataTable.id
    ).join(
        ExperimentalGroup, DataTable.group_id == ExperimentalGroup.id
    ).filter(
        DataTableMoleculeUsage.molecule_id == id
    )
    
    if start_date:
        query = query.filter(DataTable.date >= start_date)
    if end_date:
        query = query.filter(DataTable.date <= end_date)
    
    usages = query.order_by(DataTable.date).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Usage History - {molecule.name}"
    
    # Headers
    headers = [
        'Date', 'Protocol', 'Group', 'Volume Used', 'Unit', 'Animals', 'Recorded By', 'Notes'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for row_num, (usage, datatable, group) in enumerate(usages, 2):
        ws.cell(row=row_num, column=1, value=datatable.date)
        ws.cell(row=row_num, column=2, value=datatable.protocol.name)
        ws.cell(row=row_num, column=3, value=group.name)
        ws.cell(row=row_num, column=4, value=float(usage.volume_used) if usage.volume_used else 0)
        ws.cell(row=row_num, column=5, value=molecule.unit)
        ws.cell(row=row_num, column=6, value=usage.number_of_animals)
        ws.cell(row=row_num, column=7, value=usage.recorded_by.username if usage.recorded_by else '')
        ws.cell(row=row_num, column=8, value=usage.notes or '')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"usage_history_{molecule.name}_{start_date or 'all'}_{end_date or 'all'}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@controlled_molecules_bp.route('/reporting')
@login_required
@permission_required('ControlledMolecule', 'View')
def reporting():
    """Main reporting page for controlled molecules compliance."""
    # Get filter parameters
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    molecule_id = request.args.get('molecule_id', type=int)
    responsible_id = request.args.get('responsible_id', type=int)
    
    # Build query
    query = db.session.query(
        DataTableMoleculeUsage,
        DataTable,
        ExperimentalGroup,
        ControlledMolecule
    ).join(
        DataTable, DataTableMoleculeUsage.data_table_id == DataTable.id
    ).join(
        ExperimentalGroup, DataTable.group_id == ExperimentalGroup.id
    ).join(
        ControlledMolecule, DataTableMoleculeUsage.molecule_id == ControlledMolecule.id
    )
    
    # Apply filters
    if start_date:
        query = query.filter(DataTable.date >= start_date)
    if end_date:
        query = query.filter(DataTable.date <= end_date)
    if molecule_id:
        query = query.filter(DataTableMoleculeUsage.molecule_id == molecule_id)
    if responsible_id:
        query = query.filter(ControlledMolecule.responsible_id == responsible_id)
    
    results = query.order_by(DataTable.date.desc()).all()
    
    # Get filter choices
    molecules = ControlledMolecule.query.filter_by(is_active=True).order_by(ControlledMolecule.name).all()
    from app.models import User
    responsibles = db.session.query(User).join(
        ControlledMolecule, ControlledMolecule.responsible_id == User.id
    ).distinct().order_by(User.username).all()
    
    return render_template('controlled_molecules/reporting.html',
                          results=results,
                          molecules=molecules,
                          responsibles=responsibles,
                          start_date=start_date,
                          end_date=end_date,
                          molecule_id=molecule_id,
                          responsible_id=responsible_id)


@controlled_molecules_bp.route('/reporting/export')
@login_required
@permission_required('ControlledMolecule', 'View')
def export_report():
    """Export compliance report to Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    
    # Get same filters as reporting page
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    molecule_id = request.args.get('molecule_id', type=int)
    responsible_id = request.args.get('responsible_id', type=int)
    
    # Build query (same as reporting)
    query = db.session.query(
        DataTableMoleculeUsage,
        DataTable,
        ExperimentalGroup,
        ControlledMolecule
    ).join(
        DataTable, DataTableMoleculeUsage.data_table_id == DataTable.id
    ).join(
        ExperimentalGroup, DataTable.group_id == ExperimentalGroup.id
    ).join(
        ControlledMolecule, DataTableMoleculeUsage.molecule_id == ControlledMolecule.id
    )
    
    if start_date:
        query = query.filter(DataTable.date >= start_date)
    if end_date:
        query = query.filter(DataTable.date <= end_date)
    if molecule_id:
        query = query.filter(DataTableMoleculeUsage.molecule_id == molecule_id)
    if responsible_id:
        query = query.filter(ControlledMolecule.responsible_id == responsible_id)
    
    results = query.order_by(DataTable.date).all()
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Controlled Molecules Usage"
    
    # Headers
    headers = [
        'Date', 'Molecule Name', 'Regulation Category', 'Volume Used', 'Unit',
        'Number of Animals', 'Batch Number', 'Administration Route',
        'Experimental Group ID', 'Group Name', 'Project ID',
        'Responsible Person', 'Recorded By', 'Recorded At', 'Notes'
    ]
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Data rows
    for row_num, (usage, datatable, group, molecule) in enumerate(results, 2):
        ws.cell(row=row_num, column=1, value=datatable.date)
        ws.cell(row=row_num, column=2, value=molecule.name)
        ws.cell(row=row_num, column=3, value=molecule.regulation_category.value)
        ws.cell(row=row_num, column=4, value=float(usage.volume_used) if usage.volume_used else 0)
        ws.cell(row=row_num, column=5, value=molecule.unit)
        ws.cell(row=row_num, column=6, value=usage.number_of_animals)
        ws.cell(row=row_num, column=7, value=usage.batch_number or '')
        ws.cell(row=row_num, column=8, value=usage.administration_route or '')
        ws.cell(row=row_num, column=9, value=group.id)
        ws.cell(row=row_num, column=10, value=group.name)
        ws.cell(row=row_num, column=11, value=group.project_id)
        ws.cell(row=row_num, column=12, value=molecule.responsible.username if molecule.responsible else '')
        ws.cell(row=row_num, column=13, value=usage.recorded_by.username if usage.recorded_by else '')
        ws.cell(row=row_num, column=14, value=usage.recorded_at.strftime('%Y-%m-%d %H:%M') if usage.recorded_at else '')
        ws.cell(row=row_num, column=15, value=usage.notes or '')
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"controlled_molecules_report_{start_date or 'all'}_{end_date or 'all'}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
