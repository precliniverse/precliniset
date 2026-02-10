# app/helpers.py
from datetime import datetime
from io import BytesIO
import math
import numpy as np

import openpyxl
from flask import current_app, url_for, flash, has_request_context
from flask_babel import lazy_gettext as _l
from itsdangerous import URLSafeTimedSerializer
from sqlalchemy.orm.attributes import flag_modified

from .models import AnalyteDataType
from .tasks import send_email_task

try:
    from babel.support import LazyString
except ImportError:
    LazyString = None

# --- Token Generation/Verification ---
def generate_confirmation_token(email):
    """Generates a secure token for email confirmation."""
    serializer = URLSafeTimedSerializer(current_app.config['SECRET_KEY'])
    return serializer.dumps(email, salt=current_app.config['SECURITY_PASSWORD_SALT'])

def confirm_token(token, expiration=3600):
    """Confirms a token and returns the email, or None if invalid/expired."""
    serializer = URLSafeTimedSerializer(current_app.config['SECRET_KEY'])
    try:
        email = serializer.loads(
            token,
            salt=current_app.config['SECURITY_PASSWORD_SALT'],
            max_age=expiration
        )
    except Exception as e:
        current_app.logger.warning(f"Token confirmation error: {e}")
        return None
    return email

# --- Email Sending ---
def is_smtp_configured():
    """Checks if MAIL_SERVER is configured."""
    host = current_app.config.get('MAIL_SERVER')
    return bool(host and host.strip())

def send_email(to, subject, template_path, **kwargs):
    """
    Sends an email asynchronously using a Celery task.
    Returns True if the task was queued, False if SMTP is not configured.
    """
    if not is_smtp_configured():
        current_app.logger.warning(f"SMTP is not configured. Email to {to} was not sent.")
        if has_request_context():
            flash(_l("SMTP is not configured. Email to %(to)s was not sent.", to=to), 'warning')
        return False
    if 'user' in kwargs:
        from .models import User
        user = kwargs.pop('user')
        if isinstance(user, User):
            kwargs['user_id'] = user.id
        else:
            kwargs['user_id'] = None

    if LazyString and isinstance(subject, LazyString):
        subject = str(subject)

    send_email_task.delay(to, subject, template_path, **kwargs)
    return True


def sort_analytes_list_by_name(analytes_list):
    """Sorts a list of Analyte objects."""
    if not analytes_list:
        return []
    order = {'uid': 0, 'date_of_birth': 1}
    next_order_val = len(order)
    return sorted(analytes_list, key=lambda a: (order.get(a.name, next_order_val), a.name))

def get_ordered_analytes_for_model(model_id):
    """Retrieves and sorts a list of Analyte objects for a given AnimalModel ID."""
    from .extensions import db
    from .models import Analyte, AnimalModel, AnimalModelAnalyteAssociation

    model = db.session.get(AnimalModel, model_id)
    if not model:
        return []

    associations = AnimalModelAnalyteAssociation.query.filter_by(
        animal_model_id=model.id
    ).order_by(AnimalModelAnalyteAssociation.order).all()
    
    model_analytes_ordered_by_association = [assoc.analyte for assoc in associations]
    final_ordered_analytes = []
    seen_analyte_names = set()

    # Add association-defined analytes
    for analyte in model_analytes_ordered_by_association:
        if analyte.name not in seen_analyte_names:
            final_ordered_analytes.append(analyte)
            seen_analyte_names.add(analyte.name)
            
    # Add any other analytes from the model not already included
    for analyte in model.analytes:
        low_name = analyte.name.lower()
        if low_name not in seen_analyte_names:
            final_ordered_analytes.append(analyte)
            seen_analyte_names.add(low_name)

    return final_ordered_analytes

def get_ordered_columns_for_single_datatable_download(animal_model_field_defs, protocol_model_field_defs, include_age_column=True):
    """Determines the column order for downloading a single DataTable."""
    ordered_columns = []
    seen_columns = set()

    animal_fields_temp_ordered = []
    
    # Always include display_id as the primary identifier
    ordered_columns.append('display_id')
    seen_columns.add('display_id')

    if animal_model_field_defs:
        for field_def in animal_model_field_defs:
            field_name = None
            if hasattr(field_def, 'name'):
                field_name = field_def.name
            elif isinstance(field_def, (list, tuple)) and len(field_def) > 0:
                field_name = field_def[0]
            
            if field_name:
                if field_name not in {'uid', 'id', 'ID', 'display_id'}: # Exclude technical UIDs and already added display_id
                    animal_fields_temp_ordered.append(field_name)

    for field_name in animal_fields_temp_ordered:
        if field_name not in seen_columns:
            ordered_columns.append(field_name)
            seen_columns.add(field_name)

    if include_age_column and 'age_days' not in seen_columns:
        ordered_columns.append('age_days')
        seen_columns.add('age_days')

    if protocol_model_field_defs:
        for field_def in protocol_model_field_defs:
            field_name = None
            if hasattr(field_def, 'name'):
                field_name = field_def.name
            elif isinstance(field_def, (list, tuple)) and len(field_def) > 0:
                field_name = field_def[0]

            if field_name:
                if field_name not in seen_columns:
                    ordered_columns.append(field_name)
                    seen_columns.add(field_name)
    
    # Ensure uid is present in background for internal use if not explicitly added
    if 'uid' not in seen_columns:
        ordered_columns.append('uid')
        seen_columns.add('uid')

    return ordered_columns

def validate_and_convert(value, analyte_obj, field_name, row_identifier=None):
    """Validates and converts a value based on the Analyte object's data type."""
    from .models import AnalyteDataType

    if value is None or str(value).strip() == '' or str(value).strip().lower() == 'none':
        return None

    original_value = value
    value_str = str(value).strip()

    # Build location context for error messages
    location_context = ""
    if row_identifier is not None:
        location_context = f" in row {row_identifier + 1}"

    try:
        if analyte_obj.data_type == AnalyteDataType.INT:
            return int(value_str)
        elif analyte_obj.data_type == AnalyteDataType.FLOAT:
            return float(value_str)
        elif analyte_obj.data_type == AnalyteDataType.DATE:
            try:
                datetime.strptime(value_str, '%Y-%m-%d')
                return value_str
            except ValueError:
                try:
                    dt = datetime.strptime(value_str, '%Y-%m-%d %H:%M:%S')
                    return dt.strftime('%Y-%m-%d')
                except ValueError:
                    raise ValueError(
                        f"Invalid date format '{value_str}' for field '{field_name}' "
                        f"(expected YYYY-MM-DD or YYYY-MM-DD HH:MM:SS){location_context}"
                    )
        elif analyte_obj.data_type == AnalyteDataType.TEXT:
            return value_str
        elif analyte_obj.data_type == AnalyteDataType.CATEGORY:
            if analyte_obj.allowed_values:
                allowed_list = [v.strip() for v in analyte_obj.allowed_values.split(';')]
                if value_str not in allowed_list:
                    raise ValueError(
                        f"Value '{original_value}' for field '{field_name}' is not in the allowed list: "
                        f"{ ', '.join(allowed_list) }{location_context}"
                    )
            return value_str
        else:
            raise ValueError(f"Unknown analyte data type '{analyte_obj.data_type.value}' for field '{field_name}'")
    except (ValueError, TypeError) as e:
        raise ValueError(f"Invalid value '{original_value}' for field '{field_name}' (expected {analyte_obj.data_type.value}){location_context}. Error: {e}")

def generate_unique_name(base_name, existing_names_query):
    """Generates a unique name based on existing names from a query."""
    existing_names = {item.name for item in existing_names_query.all()}
    if base_name not in existing_names:
        return base_name
    count = 1
    new_name = f"{base_name}_{count}"
    while new_name in existing_names:
        count += 1
        new_name = f"{base_name}_{count}"
    return new_name

def get_ordered_column_names(data_table):
    """Gets ordered column names (Animal analytes + Protocol analytes) for a DataTable."""
    from .models import AnimalModelAnalyteAssociation, ProtocolAnalyteAssociation

    animal_analytes_ordered = []
    if data_table.group and data_table.group.model:
        associations = AnimalModelAnalyteAssociation.query.filter_by(
            animal_model_id=data_table.group.model.id
        ).order_by(AnimalModelAnalyteAssociation.order).all()
        animal_analytes_ordered = [assoc.analyte for assoc in associations]

    protocol_analytes_ordered = []
    if data_table.protocol:
        associations = ProtocolAnalyteAssociation.query.filter_by(
            protocol_model_id=data_table.protocol.id
        ).order_by(ProtocolAnalyteAssociation.order).all()
        protocol_analytes_ordered = [assoc.analyte for assoc in associations]

    ordered_columns = ['ID']
    seen_columns = {'id'}

    for analyte in animal_analytes_ordered:
        low_name = analyte.name.lower()
        col_name = 'ID' if low_name in {'animal id', 'id', 'uid', 'display_id'} else analyte.name
        if col_name not in seen_columns:
            ordered_columns.append(col_name)
            seen_columns.add(col_name)

    if 'age_days' not in seen_columns:
        ordered_columns.append('age_days')
        seen_columns.add('age_days')

    for analyte in protocol_analytes_ordered:
        low_name = analyte.name.lower()
        col_name = 'ID' if low_name in {'animal id', 'id', 'uid', 'display_id'} else analyte.name
        if col_name not in seen_columns:
            ordered_columns.append(col_name)
            seen_columns.add(col_name)

    if data_table.housing_condition:
        housing_items_ordered = sorted(data_table.housing_condition.item_associations, key=lambda x: x.item.name)
        for item_assoc in housing_items_ordered:
            if item_assoc.item.name not in seen_columns:
                ordered_columns.append(item_assoc.item.name)
                seen_columns.add(item_assoc.item.name)

    return ordered_columns

def get_field_types(data_table):
    """ Helper to get a combined dictionary of field names and their types for a DataTable. """
    from .models import AnimalModelAnalyteAssociation, ProtocolAnalyteAssociation

    field_types = {}
    
    if data_table.group and data_table.group.model:
        animal_associations = AnimalModelAnalyteAssociation.query.filter_by(
            animal_model_id=data_table.group.model.id
        ).order_by(AnimalModelAnalyteAssociation.order).all()
        for assoc in animal_associations:
            field_types[assoc.analyte.name] = assoc.analyte.data_type.value

    if data_table.protocol:
        protocol_associations = ProtocolAnalyteAssociation.query.filter_by(
            protocol_model_id=data_table.protocol.id
        ).order_by(ProtocolAnalyteAssociation.order).all()
        for assoc in protocol_associations:
            field_types[assoc.analyte.name] = assoc.analyte.data_type.value
            
    return field_types



def generate_xlsx_template(analytes, base_fields=None):
    """Generates an XLSX template file in memory."""
    from openpyxl.comments import Comment
    from openpyxl.worksheet.datavalidation import DataValidation
    from .models import AnalyteDataType

    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Header logic
    headers = ['uid', 'ID'] # uid est technique, ID est métier
    seen_headers = {'uid', 'id', 'display_id'}
    
    source_fields = base_fields or [a.name for a in analytes]
    for f in source_fields:
        low_f = f.lower()
        if low_f not in seen_headers:
            headers.append(f)
            seen_headers.add(low_f)
    
    ws.append(headers)
    
    # On masque la colonne A (uid) pour ne pas perturber l'utilisateur
    ws.column_dimensions['A'].hidden = True

    # Create a mapping for quick lookup during validation setup
    # We map both 'ID' and 'uid' to the analyte if it exists
    analyte_map = {a.name: a for a in analytes}
    if 'uid' in analyte_map:
        analyte_map['ID'] = analyte_map['uid']

    for col_idx, field_name in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        header_cell = ws.cell(row=1, column=col_idx)
        
        analyte = analyte_map.get(field_name)
        if not analyte:
            continue

        comment_lines = []
        if analyte.description:
            comment_lines.append(f"Description: {analyte.description}")
        if analyte.unit:
            comment_lines.append(f"Unit: {analyte.unit}")

        if analyte.data_type == AnalyteDataType.CATEGORY and analyte.allowed_values:
            allowed_list = [v.strip() for v in analyte.allowed_values.split(';') if v.strip()]
            # FIX: openpyxl wants a comma regardless of Excel locale
            formula = f'"{",".join(allowed_list)}"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'{col_letter}2:{col_letter}1048576')
            comment_lines.append(f"Allowed values: {', '.join(allowed_list)}")

        elif analyte.data_type == AnalyteDataType.DATE:
            dv = DataValidation(type="date", operator="greaterThan", formula1="1900-01-01")
            ws.add_data_validation(dv)
            dv.add(f'{col_letter}2:{col_letter}1048576')
            comment_lines.append("Format: YYYY-MM-DD")
            for cell in ws[col_letter][1:]:
                cell.number_format = 'YYYY-MM-DD'

        if comment_lines:
            header_cell.comment = Comment("\n".join(comment_lines), "System")

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

def create_super_admin(app, db):
    """Checks environment variables and creates the super admin if not exists."""
    from .models import User

    super_admin_email = app.config.get('SUPERADMIN_EMAIL') or "admin@example.com"
    super_admin_password = app.config.get('SUPERADMIN_PASSWORD') or "password"

    if not super_admin_email or not super_admin_password:
        return

    with app.app_context():
        try:
            # Check if User table exists by attempting a query
            # We use a try-except block for ProgrammingError (MySQL) or OperationalError (SQLite)
            # just in case the table doesn't exist yet (e.g., prior to migration).
            existing_super_admin = User.query.filter_by(is_super_admin=True).first()
            if existing_super_admin:
                return

            user = User.query.filter_by(email=super_admin_email).first()
            if user:
                if not user.is_super_admin:
                    user.is_super_admin = True
                    user.set_password(super_admin_password)
                    user.email_confirmed = True
                    db.session.add(user)
                    db.session.commit()
                return

            new_super_admin = User(
                email=super_admin_email,
                is_super_admin=True,
                email_confirmed=True
            )
            new_super_admin.set_password(super_admin_password)
            db.session.add(new_super_admin)
            db.session.commit()
        except Exception as e:
            # Likely the table doesn't exist yet, which is fine during setup/init.
            app.logger.warning(f"Skipping super admin creation: {e}")
            return

def replace_undefined(obj):
    """
    Recursively cleans data structures to ensure they are JSON serializable.
    Converts tuples to lists, None/Undefined to Python None, LazyString to string,
    and handles NaN/Infinity floats.
    """
    try:
        if hasattr(obj, '__html__') or (LazyString and isinstance(obj, LazyString)):
            return str(obj)
    except:
        pass

    if isinstance(obj, tuple):
        obj = list(obj)

    if isinstance(obj, list):
        return [replace_undefined(item) for item in obj]
    
    if isinstance(obj, dict):
        return {k: replace_undefined(v) for k, v in obj.items()}

    if obj is None:
        return None
    
    # Handle Floats (Standard and Numpy)
    if isinstance(obj, float) or isinstance(obj, np.floating):
        if math.isnan(obj) or math.isinf(obj) or np.isnan(obj) or np.isinf(obj):
            return None
        return float(obj)
    
    # Handle Numpy Integers
    if isinstance(obj, np.integer):
        return int(obj)

    if isinstance(obj, (str, int, bool)):
        return obj

    try:
        return str(obj)
    except:
        return None

def generate_display_id(group, parent_sample=None):
    """Generates a new display ID for a sample."""
    from .models import Sample

    if parent_sample:
        parent_display_id = parent_sample.display_id
        derived_count = Sample.query.filter_by(parent_sample_id=parent_sample.id).count()
        new_derived_number = derived_count + 1
        return f"{parent_display_id}-D{new_derived_number}"
    else:
        base_id = group.project.slug if group.project and group.project.slug else group.name
        last_sample = Sample.query.filter(Sample.display_id.like(f"{base_id}-S%"))\
                                  .order_by(Sample.id.desc()).first()
        
        if last_sample and last_sample.display_id:
            try:
                parts = last_sample.display_id.split('-S')
                if len(parts) == 2:
                    last_number = int(parts[1])
                    new_number = last_number + 1
                    return f"{base_id}-S{new_number}"
            except (ValueError, IndexError):
                pass

        count = Sample.query.filter(Sample.display_id.like(f"{base_id}-S%")).count()
        return f"{base_id}-S{count + 1}"

def send_workplan_update_notification(workplan_id, user_id, comment):
    """Sends an email notification to team members about a workplan update."""
    from .extensions import db
    from .models import User, Workplan
    
    workplan = db.session.get(Workplan, workplan_id)
    user = db.session.get(User, user_id)

    if not workplan or not user or not workplan.project or not workplan.project.team:
        return

    recipients = [membership.user.email for membership in workplan.project.team.memberships]
    if not recipients:
        return

    workplan_url = url_for('workplans.edit_workplan', workplan_id=workplan.id, _external=True)
    
    send_email(
        to=recipients,
        subject=str(_l('Workplan Update: [%(project_slug)s] %(workplan_name)s', 
                 project_slug=workplan.project.slug, workplan_name=workplan.name)),
        template_path='email/workplan_update.html',
        project_id=workplan.project.id,
        workplan_id=workplan.id,
        user_id=user.id,
        comment=comment,
        workplan_url=workplan_url
    )

def clean_param_name_for_id(name):
    """Cleans a parameter name to be safe for use as an HTML ID."""
    import re
    cleaned_name = re.sub(r'[^\w-]', '-', name)
    cleaned_name = re.sub(r'-+', '-', cleaned_name)
    cleaned_name = cleaned_name.strip('-')
    if not cleaned_name:
        return "invalid-name"
    return cleaned_name

def ensure_mandatory_analytes_exist(app, db):
    """Ensures that the mandatory 'uid' and 'date_of_birth' analytes exist."""
    with app.app_context():
        try:
            from sqlalchemy import inspect
            from .models import Analyte, AnalyteDataType

            inspector = inspect(db.engine)
            if not inspector.has_table("analyte"):
                return

            mandatory_analytes = {
                "uid": {"type": AnalyteDataType.TEXT, "description": "Unique animal identifier"},
                "date_of_birth": {"type": AnalyteDataType.DATE, "description": "Animal's date of birth"}
            }

            for name, details in mandatory_analytes.items():
                analyte = Analyte.query.filter_by(name=name).first()
                if not analyte:
                    new_analyte = Analyte(
                        name=name,
                        data_type=details["type"],
                        description=details["description"],
                        is_mandatory=True
                    )
                    db.session.add(new_analyte)
                elif not analyte.is_mandatory:
                    analyte.is_mandatory = True
            
            db.session.commit()
        except Exception as e:
            app.logger.warning(f"Skipping mandatory analyte check: {e}")
            pass

def get_project_or_404(project_slug_or_id, permission_needed='read'):
    """Fetches a project by its slug or ID and checks user permissions."""
    from werkzeug.exceptions import Forbidden, NotFound
    from .extensions import db
    from .models import Project
    from .permissions import check_project_permission
    
    project = None
    if str(project_slug_or_id).isdigit():
        project = db.session.get(Project, int(project_slug_or_id))
    if not project:
        project = Project.query.filter_by(slug=project_slug_or_id).first()
    
    if not project:
        raise NotFound(f"Project with identifier '{project_slug_or_id}' not found.")
    
    if not check_project_permission(project, permission_needed): 
         raise Forbidden(f"You do not have '{permission_needed}' permission for this project.")
    return project